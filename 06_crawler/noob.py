import requests
from lxml import etree
from openpyxl import Workbook


def main():
    # 网页url
    url = 'https://www.runoob.com/html/html-tutorial.html'
    ua_headers = {"User-Agent": 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)'}
    # 网页代码
    response = requests.get(url=url, headers=ua_headers).text
    # 转换为etree对象
    tree = etree.HTML(response)

    # 匹配到所有 target 属性为 _top 的 a 标签,返回一个列表
    result_lst = tree.xpath('//a[@target="_top"]')
    wb = Workbook()
    ws = wb.active
    for i, a in enumerate(result_lst, start=1):
        print(a.attrib)
        title = a.attrib.get("title")
        if not title:
            break
        ws.cell(i, 1, title)
        ws.cell(i, 2, a.attrib.get("href"))
    wb.save('noob.xlsx')


if __name__ == '__main__':
    main()
